#!/usr/bin/env python
"""
Load data from the spreadsheet to the database.
"""

import argparse
from datetime import datetime, timedelta
import glob
import logging
import os
import subprocess
import sqlite3
from openpyxl import load_workbook
import re


log = logging.getLogger(__name__)


def build_parser():
    parser = argparse.ArgumentParser(description = globals()['__doc__'],
                                     formatter_class = argparse.RawDescriptionHelpFormatter)
    parser.add_argument('-f',
                        '--in_file',
                        default='investments.xlsx',
                        help='The spreadsheet input file. '
                        'Default: %(default)s.')
    parser.add_argument('-d',
                        '--db_file',
                        default='investments.db',
                        help='Sqlite database filename. '
                        'Default: %(default)s.')
    return parser


def run_date(arg):
    popen_args = ['date', '-d', arg, '+%s']
    proc = subprocess.Popen(popen_args, stdout=subprocess.PIPE)
    return  proc.stdout.read()


class App(object):
    def __init__(self, args):
        self.args = args
        self.db_open()
        self.db_init()
        self.wb = load_workbook(filename = self.args.in_file)


    def db_open(self):
        self.con = sqlite3.connect(self.args.db_file)
        self.cur = self.con.cursor()


    def db_commit(self):
        self.con.commit()


    def db_close(self):
        self.con.close()


    def db_init(self):
        # Set up these three tables.
        self.init_account()
        self.init_performance_review()
        self.init_trade_confirmation()
        self.init_activity()
        self.init_trade_history()


    def eval(self, ws, in_expr):
        """
        Try to evaluate a cell expression. Assumed to start with an '='.
        """
        # change cell references to: ws["cellname"].value
        if isinstance(in_expr, int):
            return in_expr
        expr = re.sub(r'([A-Z]\d+)', r'ws["\1"].value', in_expr[1:])
        return eval(expr)


    def init_account(self):
        sql = 'DROP TABLE IF EXISTS account'
        self.cur.execute(sql)
        sql = """CREATE TABLE IF NOT EXISTS account(
  number text,
  name text
)"""
        self.cur.execute(sql)


    def init_performance_review(self):
        sql = 'DROP TABLE IF EXISTS performance_review'
        self.cur.execute(sql)
        sql = """CREATE TABLE IF NOT EXISTS performance_review (
	id INTEGER PRIMARY KEY AUTOINCREMENT,
	end_date TEXT,
	account TEXT,
	end_market_value REAL,
	gain REAL
)"""
        self.cur.execute(sql)


    def init_trade_confirmation(self):
        sql = 'DROP TABLE IF EXISTS trade_confirmation'
        self.cur.execute(sql)
        sql = """CREATE TABLE IF NOT EXISTS trade_confirmation(
  id integer PRIMARY KEY AUTOINCREMENT,
  symbol text,
  trade_date text,
  is_buy integer,
  n_shares integer,
  share_price real,
  total real,
  account text,
  fee real DEFAULT 0.0,
  accrued_interest real DEFAULT 0.0,
  trade_type text,
  name text,
  expiration_date text DEFAULT '',
  strike_price real DEFAULT 0.0
)"""
        # is_buy - 1=buy, 0=sell
        # trade_type - stock (default), call, bond, preferred stock
        # expiration_date - only for options, NULL otherwise
        # strike_price - only for options, NULL otherwise
        self.cur.execute(sql)


    def init_activity(self):
        sql = 'DROP TABLE IF EXISTS activity'
        self.cur.execute(sql)
        sql = """CREATE TABLE IF NOT EXISTS activity(
  id integer PRIMARY KEY AUTOINCREMENT,
  account text,
  activity_date text,
  amount real,
  name text,
  symbol text,
  n_shares integer,
  activity_type text)
"""
        # activity_type - dividend, purchase, sale, interest, fee, bought
        self.cur.execute(sql)


    def init_trade_history(self):
        sql = 'DROP TABLE IF EXISTS trade_history'
        self.cur.execute(sql)
        sql = """CREATE TABLE IF NOT EXISTS trade_history(
  id integer PRIMARY KEY AUTOINCREMENT,
  account text,
  history_date text,
  symbol text,
  n_shares integer,
  unit_cost real,
  current_price real,
  name text)
"""
        self.cur.execute(sql)


    def load_accounts(self):
        start_row = 3
        first_row = True
        # Grab the account details from this tab.
        # There isn't a tab specifically for this.
        ws = self.wb['Trade Confirmations']
        count = 0
        for row in ws.iter_rows(min_row=start_row):
            if first_row:
                numbers = [cell.value for cell in row[1:7] if cell.value]
                first_row = False
            else:
                names = [cell.value for cell in row[1:7] if cell.value]
                break
        sql = "INSERT INTO account(number, name) values(:number, :name)"
        for (number, name) in zip(numbers, names):
            count += 1
            self.cur.execute(sql, dict(number=number, name=name))
        print('accounts:', count)


    def load_performance_reviews(self):
        """
        Get selected detail from the Perf Reviews tab.
        Ignore the total amounts since they can be calculated.
        """
        start_row = 7           # Row numbers are 1 based
        ws = self.wb['Perf Reviews']
        count = 0
        col_names = ['end_date', 'account', 'end_market_value', 'gain']
        accounts = ['5304-3149', '4796-5300', '3029-7830']
        # Need to manually calculate the gain.
        end_market_value_cols = [6, 8, 10] # col numbers are 0 based
        prev_end_market_values = [0] * len(accounts)
        sql = ("INSERT INTO performance_review(" + ','.join(col_names) + ")" 
               "values(" + ','.join([':' + col for col in col_names]) + ")")
        for row in ws.iter_rows(min_row=start_row):
            end_date = row[0].value
            if not end_date:
                break
            all_cols = list(zip(accounts, end_market_value_cols, prev_end_market_values))
            for (account, end_market_value_col, prev_end_market_value) in all_cols:
                if prev_end_market_value:
                    gain = row[end_market_value_col].value - prev_end_market_value
                else:
                    gain = 0.0
                values = dict(end_date=end_date,
                              account=account,
                              end_market_value=row[end_market_value_col].value,
                              gain=gain)
                count += 1
                self.cur.execute(sql, values)
            prev_end_market_values = [row[col].value for col in end_market_value_cols]
        print('performance_reviews:', count)


    def load_trade_confirmations(self):
        start_row = 6
        ws = self.wb['Trade Confirmations']
        first_row = True
        count = 0
        for row in ws.iter_rows(min_row=start_row):
            if first_row:
                first_row = False
                columns = [cell.value for cell in row if cell.value is not None]
            else:
                values= dict(list(zip(columns, [cell.value for cell in row])))
                if not values['account']:
                    break
                count += 1
                if values['account'].startswith('='):
                    # The acount columns point to an earlier cell that has the value.
                    values['account'] = ws[values['account'][1:]].value
                values['total'] = self.eval(ws, values['total'])
                for (look_for, trade_type) in [(' bonds', 'bond'), (' preferred', 'preferred stock')]:
                    if values['name'].endswith(look_for):
                        # Leave the name alone for now. Seems to match better this way.
                        #values['name'] = values['name'][:-len(look_for)]
                        values['trade_type'] = trade_type
                # Look for detail that isn't a comment.
                if values['detail'] and values['detail'][0] != '#':
                    for pair in values['detail'].split(','):
                        (name, value) = pair.split('=', 1)
                        if name.endswith('_date'):
                            values[name] = datetime.strptime(value, '%m/%d/%Y')
                        elif value.lower() in ['true', 'yes']:
                            values[name] = True
                        else:
                            values[name] = value
                # Fill in default values:
                for (name, value) in [('expiration_date', None), ('strike_price', None),
                                      ('trade_type', 'stock'), ('fee', None),
                                      ('accrued_interest', None)]:
                    if name not in values:
                        values[name] = value
                sql = """INSERT INTO trade_confirmation (trade_date, is_buy, n_shares, share_price, total, account,
  fee, accrued_interest, trade_type, symbol, name, expiration_date, strike_price)
VALUES(:trade_date, :is_buy, :n_shares, :share_price, :total, :account,
  :fee, :accrued_interest, :trade_type, :symbol, :name, :expiration_date, :strike_price)
"""
                self.cur.execute(sql, values)
        print('trade confirmations:', count)


    def load_account_detail(self):
        for account in ['Individual', 'Roth IRA', 'IRA']:
            self.load_activity(account)
            self.load_trade_history(account)


    def load_activity(self, sheet_name):
        start_row = 5
        row_number = start_row
        ws = self.wb[sheet_name]
        account = ws['B1'].value
        count = 0
        try:
            for row in ws.iter_rows(min_row=start_row):
                row_number += 1
                activity_date = row[0].value
                if activity_date == 'END':
                    # The last line is marked.
                    break
                if not activity_date:
                    # Skip empty rows.
                    continue
                count += 1
                amount = float(row[1].value)
                symbol = row[2].value
                name = row[3].value
                n_shares = None
                activity_type = None
                for activity_type_re in [r'\s+(dividend)$',
                                         r'\s+(purchase)$',
                                         r'\s+(sale)$',
                                         r'\s+(interest)$',
                                         r'\s+(cash)$',
                                         r'\s+(withholding)$',
                                         r'\s+(name\s+change)$',
                                         r'\s+(short\s+term\s+cap\s+gain)$',
                                         r'\s+(call\s+(?:assigned|expired))$',
                                         r'\s+(dividend\s+withholding)$',
                                         r'\s+(dividend\s+reinvestment)$',
                                         r'(Pass\s+thru\s+to\s+Roth\s+IRA\s+from\s+individual)$',
                                         #    11111   222222222222222222222222222222222         33333             4444444444444
                                         r'\s+(\d+)\s+(calls\s+(?:bought|expires|sold))\s+@\s+\$(\d+)\s+expires\s+(\d+/\d+/\d+)$',
                                         r'\s+(call\s+(?:bought|sold))$',
                                         r'\s+(stock\s+distribution)$',
                                         r'\s+(redemption)$',
                                         r'\s+(fee)$',
                                         r'\s+(Reimburse MF fees\s+\dQ\d\d)$',
                                         r'\s+(bought|sold)\s+(\d+)(?:\.\d+)?$']:
                    m = re.search(activity_type_re, name)
                    if m:
                        activity_type = m.group(1)
                        if len(m.groups()) == 2:
                            n_shares = int(m.group(2))
                        elif len(m.groups()) == 4:
                            activity_type = m.group(2)
                            # TODO:
                            # What do I do with the #calls, strike price and expiration date?
                        name = name[:m.start()]
                values = dict(account = account,
                              activity_date = activity_date,
                              amount = amount,
                              name = name,
                              symbol = symbol,
                              n_shares = n_shares,
                              activity_type = activity_type)
                if activity_type is None:
                    print('Error: Unable to determine activity_type at row %s for %s' % (row_number, values))
                    print('description:', name, ' sheet:', sheet_name)
                    exit()
                sql = """INSERT INTO activity (account, activity_date, amount, name, symbol, n_shares, activity_type)
VALUES(:account, :activity_date, :amount, :name, :symbol, :n_shares, :activity_type)
"""
                self.cur.execute(sql, values)
            print('%s activities for %s: %d' % (sheet_name, account, count))
        except Exception as e:
            print('Exception %s at row %d' % (str(e), row_number))
            raise e


    def load_trade_history(self, sheet_name):
        start_row = 5
        row_number = start_row
        ws = self.wb[sheet_name]
        account = ws['B1'].value
        count = 0
        try:
            for row in ws.iter_rows(min_row=start_row):
                row_number += 1
                if row[0] == 'END':
                    # The last line is marked.
                    break
                values = dict(list(zip(
                    # 6               7         8           9            10               13
                    ['history_date', 'symbol', 'n_shares', 'unit_cost', 'current_price', 'name'], 
                    [cell.value for cell in row[6:11] + (row[13],)])))
                if not values['symbol']:
                    # Skip empty rows.
                    continue
                # Fill in values that may not be in this row
                if values['history_date']:
                    history_date = values['history_date']
                else:
                    values['history_date'] = history_date
                values['account'] = account
                count += 1
                sql = """INSERT INTO trade_history (account, history_date, symbol, n_shares, unit_cost, current_price, name)
VALUES(:account, :history_date, :symbol, :n_shares, :unit_cost, :current_price, :name)
"""
                self.cur.execute(sql, values)
                              
        except Exception as e:
            print('Exception %s at row %d' % (str(e), row_number))
            print(values)
            raise e
        print('trade history for %s: %d' % (account, count))

        
def action(args):
    """
    Load the parts in the spreadsheet
    """
    app = App(args)
    app.load_accounts()
    app.load_performance_reviews()
    app.load_trade_confirmations()
    app.load_account_detail()
    app.db_commit()
    app.db_close()
    

if __name__ == '__main__':
    action(build_parser().parse_args())
